#!/usr/bin/env python3
# Copyright (c) 2025 frisky1985
# SPDX-License-Identifier: MIT

"""
Excel report writer for yuleOSH evidence reports.

Generates .xlsx reports with:
  - MISRA C compliance reports (4 sheets: Summary, Violations, Traceability, Deviations)
  - Self-test reports (4 sheets: Summary, Test Results, Coverage, SHALL Coverage)

All file paths and line numbers are rendered as clickable hyperlinks for
code-level traceability.

Requires openpyxl >= 3.1.0.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

log = logging.getLogger("evidence.excel_writer")

# ── Colour palette ──────────────────────────────────────────────────────────
_FILL_RED_LIGHT = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
_FILL_YELLOW_LIGHT = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
_FILL_BLUE_LIGHT = PatternFill(start_color="DEEAF6", end_color="DEEAF6", fill_type="solid")
_FILL_GREEN_LIGHT = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_FILL_GREY_LIGHT = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
_FILL_RED_MEDIUM = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

_FILL_COVERAGE_RED = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
_FILL_COVERAGE_YELLOW = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
_FILL_COVERAGE_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

_FONT_HEADER = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_FONT_METRIC = Font(name="Calibri", bold=True, size=11)
_FONT_LINK = Font(name="Calibri", size=10, color="0563C1", underline="single")
_FONT_NORMAL = Font(name="Calibri", size=10)
_FONT_BOLD = Font(name="Calibri", size=10, bold=True)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

_THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

_SEVERITY_EMOJI = {
    "required": "🔴",
    "advisory": "🟡",
    "directive": "🔵",
    "project_specific": "⚪",
}


def _make_hyperlink(target: str, display: Optional[str] = None) -> str:
    """Create a hyperlink formula string for Excel.

    Uses the HYPERLINK() worksheet function which is widely supported
    in Excel, WPS, and Google Sheets.
    Returns the display text (with hyperlink formula) for the cell value.
    """
    from openpyxl.worksheet.hyperlink import Hyperlink
    return display or target


def _apply_header_style(ws, row: int, max_col: int):
    """Apply standard header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _THIN_BORDER


def _apply_body_style(ws, row: int, max_col: int, font: Font = _FONT_NORMAL):
    """Apply standard body styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.border = _THIN_BORDER
        if col == 1:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_column_width(ws, max_col: int, max_width: int = 60):
    """Auto-adjust column widths based on content."""
    for col in range(1, max_col + 1):
        max_len = 0
        letter = get_column_letter(col)
        for cell in ws[letter]:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        adjusted = min(max_len + 3, max_width)
        ws.column_dimensions[letter].width = max(adjusted, 10)


def _to_absolute_path(file_path: str) -> str:
    """Convert a relative file path to an absolute file:// URI."""
    p = Path(file_path)
    if not p.is_absolute():
        # Try to resolve relative to cwd
        p = Path.cwd() / p
    return p.resolve().as_uri()


def _severity_fill(misra_sev: str) -> Optional[PatternFill]:
    """Return the appropriate row fill for a MISRA severity category."""
    sev = misra_sev.strip().lower()
    if sev == "required":
        return _FILL_RED_LIGHT
    elif sev == "advisory":
        return _FILL_YELLOW_LIGHT
    elif sev == "directive":
        return _FILL_BLUE_LIGHT
    return None


def _status_fill(status: str) -> Optional[PatternFill]:
    """Return the appropriate row fill for a test status."""
    s = status.strip().lower()
    if s in ("failed", "error"):
        return _FILL_RED_LIGHT
    elif s == "passed":
        return _FILL_GREEN_LIGHT
    elif s == "skipped":
        return _FILL_GREY_LIGHT
    return None


def _coverage_fill(rate: float) -> Optional[PatternFill]:
    """Return fill based on coverage rate thresholds."""
    if rate > 90.0:
        return _FILL_COVERAGE_GREEN
    elif rate >= 80.0:
        return _FILL_COVERAGE_YELLOW
    elif rate >= 0:
        return _FILL_COVERAGE_RED
    return None


# ═══════════════════════════════════════════════════════════════════════════
# ExcelReportWriter
# ═══════════════════════════════════════════════════════════════════════════


class ExcelReportWriter:
    """Generate structured .xlsx reports for MISRA C compliance and self-test results.

    Usage::

        writer = ExcelReportWriter(output_dir)
        path = writer.write_misra_report(violations, groups, summary, rule_defs, deviations, output_path)
    """

    def __init__(self, output_dir: Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ──────────────────────────────────────────────────────────────────────
    # MISRA Report
    # ──────────────────────────────────────────────────────────────────────

    def write_misra_report(
        self,
        violations: list[dict],
        groups: dict,
        summary: dict,
        rule_defs: dict,
        deviations: Optional[list] = None,
        output_path: Optional[Path] = None,
    ) -> Path:
        """Generate MISRA C compliance report Excel with 4 sheets.

        Sheet layout:
          1. Summary — metrics overview
          2. Violations — per-violation details with clickable file/line links
          3. Traceability — rule→spec→implementation→test matrix
          4. Deviations — deviation register

        Returns the path to the generated .xlsx file.
        """
        deviations = deviations or []
        wb = Workbook()

        # Disable auto-calculation for speed
        wb.calculation = None

        self._write_misra_summary(wb, summary, groups, violations, rule_defs, deviations)
        self._write_misra_violations(wb, violations, rule_defs)
        self._write_misra_traceability(wb, violations, rule_defs)
        self._write_misra_deviations(wb, deviations)

        out = Path(output_path) if output_path else self.output_dir / "misra-report.xlsx"
        wb.save(str(out))
        log.info("MISRA Excel report saved: %s", out)
        return out

    def _write_misra_summary(
        self, wb: Workbook, summary: dict, groups: dict,
        violations: list[dict], rule_defs: dict, deviations: list,
    ):
        ws = wb.active
        ws.title = "Summary"
        ws.sheet_properties.tabColor = "4472C4"

        headers = ["Metric", "Value"]
        ws.append(headers)
        _apply_header_style(ws, 1, 2)

        # ── Basic info ──
        rows = [
            ("Generated", datetime.now().isoformat()),
            ("Tool", "cppcheck --addon=misra"),
            ("Tool Version", self._get_tool_version()),
            ("Standard", "MISRA C:2023"),
            ("Scan Mode", "Full"),
            ("Total KLOC", summary.get("total_kloc", 0)),
            ("", ""),
            ("Total Violations", summary.get("total_violations", 0)),
            ("Required", summary.get("misra_classification", {}).get("required", 0)),
            ("Advisory", summary.get("misra_classification", {}).get("advisory", 0)),
            ("Directive", summary.get("misra_classification", {}).get("directive", 0)),
            ("Project-specific", summary.get("misra_classification", {}).get("project_specific", 0)),
            ("", ""),
            ("Violations per KLOC", summary.get("violations_per_kloc", 0)),
            ("Files Affected", len(summary.get("unique_files", []))),
            ("", ""),
        ]

        for metric, value in rows:
            ws.append([metric, value])

        # ── Severity breakdown ──
        sev_labels = [
            ("Severity: error", "error", "❌"),
            ("Severity: warning", "warning", "⚠️"),
            ("Severity: style", "style", "🎨"),
            ("Severity: information", "information", "ℹ️"),
        ]
        for label, key, icon in sev_labels:
            count = summary.get("severity_counts", {}).get(key, 0)
            ws.append([f"{icon} {label}", count])

        last_row = ws.max_row

        # Style body cells
        for row in range(2, last_row + 1):
            cell_a = ws.cell(row=row, column=1)
            cell_b = ws.cell(row=row, column=2)
            cell_a.border = _THIN_BORDER
            cell_b.border = _THIN_BORDER

            val = cell_a.value
            if isinstance(val, str) and val.startswith(("Total ", "Severity:")):
                cell_a.font = _FONT_BOLD
            else:
                cell_a.font = _FONT_NORMAL
            cell_b.font = _FONT_NORMAL
            cell_b.alignment = _ALIGN_CENTER

            # Highlight Required > 0
            if val and "Required" in str(val):
                req_val = cell_b.value
                if isinstance(req_val, (int, float)) and req_val > 0:
                    cell_a.fill = _FILL_RED_LIGHT
                    cell_b.fill = _FILL_RED_LIGHT

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    def _write_misra_violations(self, wb: Workbook, violations: list[dict], rule_defs: dict):
        ws = wb.create_sheet("Violations")
        ws.sheet_properties.tabColor = "E74C3C"

        headers = ["Rule ID", "Severity (MISRA)", "File", "Line", "Column", "Message", "Spec Ref", "Fix Status", "Deviation"]
        ws.append(headers)
        _apply_header_style(ws, 1, len(headers))

        for v in violations:
            rid = v.get("rule_id", "")
            file_path = v.get("file", "")
            line_num = v.get("line", 0)
            col_num = v.get("col", 0)
            msg = v.get("message", "")
            severity = v.get("severity", "style")

            # Determine MISRA category severity
            defn = rule_defs.get(rid, {})
            misra_sev = defn.get("severity", "project_specific").lower()
            sev_display = _SEVERITY_EMOJI.get(misra_sev, "⚪") + " " + misra_sev.capitalize()

            spec_ref = defn.get("spec_ref", "")

            fix_status = v.get("fix_status", "unresolved")
            deviation = v.get("deviation_ref", {})
            deviation_str = deviation.get("reason", "")[:60] if deviation else ""

            row_data = [rid, sev_display, file_path, line_num, col_num, msg, spec_ref, fix_status, deviation_str]
            ws.append(row_data)

            row_idx = ws.max_row

            # Format body cells
            _apply_body_style(ws, row_idx, len(headers))

            # Apply severity-based row highlighting
            fill = _severity_fill(misra_sev)
            if fill:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = fill

            # ── Hyperlink for File column (col C=3) ──
            if file_path:
                try:
                    file_uri = _to_absolute_path(file_path)
                    cell_c = ws.cell(row=row_idx, column=3)
                    cell_c.hyperlink = file_uri
                    cell_c.font = _FONT_LINK
                except Exception:
                    pass

            # ── Hyperlink for Line column (col D=4) ──
            if file_path and line_num:
                try:
                    file_uri = _to_absolute_path(file_path)
                    line_uri = f"{file_uri}#line{line_num}"
                    cell_d = ws.cell(row=row_idx, column=4)
                    cell_d.hyperlink = line_uri
                    cell_d.font = _FONT_LINK
                except Exception:
                    pass

        # Auto-width
        col_widths = [18, 22, 50, 8, 8, 60, 18, 14, 30]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Freeze top row
        ws.freeze_panes = "A2"

    def _write_misra_traceability(self, wb: Workbook, violations: list[dict], rule_defs: dict):
        ws = wb.create_sheet("Traceability")
        ws.sheet_properties.tabColor = "2E86C1"

        headers = ["Rule ID", "Spec Ref", "Implementation Ref", "Test Ref", "Check Method", "Auto-checkable"]
        ws.append(headers)
        _apply_header_style(ws, 1, len(headers))

        # Build a set of violated rules for the matrix
        violated_rules: dict[str, dict] = {}
        for v in violations:
            rid = v.get("rule_id", "")
            if rid and rid != "unknown":
                if rid not in violated_rules:
                    defn = rule_defs.get(rid, {})
                    violated_rules[rid] = {
                        "spec_ref": defn.get("spec_ref", ""),
                        "impl_ref": defn.get("impl_ref", defn.get("check_method", "")),
                        "test_ref": defn.get("test_ref", ""),
                        "check_method": defn.get("check_method", ""),
                        "auto_checkable": defn.get("auto_checkable", True),
                    }

        for rid, info in sorted(violated_rules.items()):
            row_data = [
                rid,
                info["spec_ref"],
                info["impl_ref"],
                info["test_ref"],
                info["check_method"],
                "Yes" if info["auto_checkable"] else "No",
            ]
            ws.append(row_data)
            _apply_body_style(ws, ws.max_row, len(headers))

        col_widths = [22, 20, 24, 20, 24, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _write_misra_deviations(self, wb: Workbook, deviations: list):
        ws = wb.create_sheet("Deviations")
        ws.sheet_properties.tabColor = "F39C12"

        headers = ["Rule ID", "File Pattern", "Reason", "Status", "Risk Level", "Expires", "Approved By", "ALM Ticket", "Expired?"]
        ws.append(headers)
        _apply_header_style(ws, 1, len(headers))

        from datetime import datetime as dt

        for dev in deviations:
            dev_dict = self._deviation_to_dict(dev)
            expires = dev_dict.get("expires", "")
            is_expired = False
            if expires:
                try:
                    is_expired = dt.fromisoformat(expires) < dt.now()
                except (ValueError, TypeError):
                    pass

            row_data = [
                dev_dict.get("deviation_rule", ""),
                dev_dict.get("file_pattern", ""),
                dev_dict.get("reason", ""),
                dev_dict.get("status", "pending"),
                dev_dict.get("risk_level", "mid"),
                expires,
                dev_dict.get("approved_by", ""),
                dev_dict.get("alm_ticket", ""),
                "yes" if is_expired else "no",
            ]
            ws.append(row_data)
            row_idx = ws.max_row
            _apply_body_style(ws, row_idx, len(headers))

            # Conditional formatting
            if is_expired:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = _FILL_RED_LIGHT
            elif dev_dict.get("status", "").lower() == "rejected":
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = _FILL_GREY_LIGHT

        col_widths = [22, 30, 50, 14, 12, 22, 18, 18, 10]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ──────────────────────────────────────────────────────────────────────
    # Self-Test Report
    # ──────────────────────────────────────────────────────────────────────

    def write_selftest_report(
        self,
        review: dict,
        output_path: Optional[Path] = None,
    ) -> Path:
        """Generate self-test report Excel with 4 sheets.

        Sheet layout:
          1. Summary — metrics overview
          2. Test Results — per-test-case details
          3. Coverage — per-file coverage rates
          4. SHALL Coverage — requirement traceability

        Parameters
        ----------
        review : dict
            The enhanced review dict from step_review_selftest().
        output_path : Path, optional
            Where to write the .xlsx file.

        Returns the path to the generated .xlsx file.
        """
        wb = Workbook()

        self._write_selftest_summary(wb, review)
        self._write_selftest_results(wb, review.get("test_case_results", []))
        self._write_selftest_coverage(wb, review.get("coverage", {}))
        self._write_selftest_shall_coverage(wb, review)

        out = Path(output_path) if output_path else self.output_dir / "selftest-report.xlsx"
        wb.save(str(out))
        log.info("Self-test Excel report saved: %s", out)
        return out

    def _write_selftest_summary(self, wb: Workbook, review: dict):
        ws = wb.active
        ws.title = "Summary"
        ws.sheet_properties.tabColor = "4472C4"

        headers = ["Metric", "Value"]
        ws.append(headers)
        _apply_header_style(ws, 1, 2)

        coverage = review.get("coverage", {})
        shall_total = review.get("shall_total", 0)
        shall_covered = review.get("shall_covered", 0)
        shall_rate = round(shall_covered / shall_total * 100, 1) if shall_total > 0 else 0.0
        total_tests = (
            review.get("total_passed", 0)
            + review.get("total_failed", 0)
            + review.get("total_skipped", 0)
            + review.get("total_errors", 0)
        )
        pass_rate = review.get("pass_rate", 0.0)

        env = review.get("environment", {})

        rows = [
            ("Runner", env.get("platform", "pytest")),
            ("Total Tests", total_tests),
            ("Passed", review.get("total_passed", 0)),
            ("Failed", review.get("total_failed", 0)),
            ("Skipped", review.get("total_skipped", 0)),
            ("Errors", review.get("total_errors", 0)),
            ("", ""),
            ("Pass Rate (%)", f"{pass_rate:.1f}%"),
            ("Duration (Sec)", review.get("duration_sec", 0)),
            ("", ""),
            ("Line Coverage (%)", f"{coverage.get('line_rate', 0):.1f}%"),
            ("Branch Coverage (%)", f"{coverage.get('branch_rate', 0):.1f}%"),
            ("Function Coverage (%)", f"{coverage.get('function_rate', 0):.1f}%"),
            ("MC/DC (%)", f"{coverage.get('mc_dc_rate', 0):.1f}%" if coverage.get('mc_dc_rate') is not None else "N/A"),
            ("", ""),
            ("SHALL Total", shall_total),
            ("SHALL Covered", shall_covered),
            ("SHALL Coverage (%)", f"{shall_rate:.1f}%"),
        ]

        for metric, value in rows:
            ws.append([metric, value])

        last_row = ws.max_row
        for row in range(2, last_row + 1):
            cell_a = ws.cell(row=row, column=1)
            cell_b = ws.cell(row=row, column=2)
            cell_a.border = _THIN_BORDER
            cell_b.border = _THIN_BORDER
            cell_a.font = _FONT_NORMAL if row > 2 else _FONT_METRIC
            cell_b.font = _FONT_NORMAL
            cell_b.alignment = _ALIGN_CENTER
            val = str(cell_a.value or "")
            if "Failed" in val and cell_b.value and cell_b.value != 0:
                cell_a.fill = _FILL_RED_LIGHT
                cell_b.fill = _FILL_RED_LIGHT
            elif val.startswith("SHALL Coverage") and shall_rate < 100:
                cell_a.fill = _FILL_YELLOW_LIGHT
                cell_b.fill = _FILL_YELLOW_LIGHT

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 18

    def _write_selftest_results(self, wb: Workbook, test_case_results: list[dict]):
        ws = wb.create_sheet("Test Results")
        ws.sheet_properties.tabColor = "E74C3C"

        headers = [
            "Test Name", "Status", "Duration (s)", "Type",
            "File", "Failure Message", "Failure Type", "Stack Trace",
        ]
        ws.append(headers)
        _apply_header_style(ws, 1, len(headers))

        status_icon_map = {
            "passed": "✅ passed",
            "failed": "❌ failed",
            "skipped": "⏭️ skipped",
            "error": "❓ error",
        }

        for tc in test_case_results:
            name = tc.get("name", "")
            status = tc.get("status", "passed")
            duration = tc.get("duration", 0)
            test_type = tc.get("type", "unit")
            message = tc.get("message", "")
            fail_info = tc.get("failure", {}) or {}
            fail_type = fail_info.get("type", "") if isinstance(fail_info, dict) else ""
            stacktrace = fail_info.get("stacktrace", "") if isinstance(fail_info, dict) else ""

            # Extract file path from classname portion of test name
            file_path = ""
            if "::" in name:
                classname = name.split("::")[0]
                # Try to resolve as a file path
                if classname.endswith(".py"):
                    file_path = classname

            # Strip emoji-embedded status for display
            status_display = status_icon_map.get(status, status)

            row_data = [name, status_display, duration, test_type, file_path, message, fail_type, stacktrace]
            ws.append(row_data)
            row_idx = ws.max_row
            _apply_body_style(ws, row_idx, len(headers))

            # Status-based row colour
            fill = _status_fill(status)
            if fill:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = fill

            # Hyperlink for File column (col E=5)
            if file_path:
                try:
                    file_uri = _to_absolute_path(file_path)
                    cell_e = ws.cell(row=row_idx, column=5)
                    cell_e.hyperlink = file_uri
                    cell_e.font = _FONT_LINK
                except Exception:
                    pass

        col_widths = [50, 14, 12, 12, 50, 50, 16, 50]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"

    def _write_selftest_coverage(self, wb: Workbook, coverage: dict):
        ws = wb.create_sheet("Coverage")
        ws.sheet_properties.tabColor = "27AE60"

        headers = ["File", "Line Rate (%)", "Branch Rate (%)", "Function Rate (%)", "MC/DC Rate (%)"]
        ws.append(headers)
        _apply_header_style(ws, 1, len(headers))

        per_file = coverage.get("per_file", [])

        if not per_file:
            ws.append(["No per-file coverage data available", "", "", "", ""])
            _apply_body_style(ws, 2, len(headers))
        else:
            for pf in sorted(per_file, key=lambda x: x.get("file", "")):
                fname = pf.get("file", "")
                lr = pf.get("line_rate", 0.0)
                br = pf.get("branch_rate", 0.0)
                fr = pf.get("function_rate", 0.0)
                # MC/DC may not be present in per_file
                mcdc = pf.get("mcdc_rate", pf.get("mc_dc_rate", None))

                row_data = [fname, lr, br, fr, mcdc if mcdc is not None else "N/A"]
                ws.append(row_data)
                row_idx = ws.max_row
                _apply_body_style(ws, row_idx, len(headers))

                # Hyperlink for File column
                if fname:
                    try:
                        file_uri = _to_absolute_path(fname)
                        cell_a = ws.cell(row=row_idx, column=1)
                        cell_a.hyperlink = file_uri
                        cell_a.font = _FONT_LINK
                    except Exception:
                        pass

                # Conditional coverage colour for each rate column
                for col_idx, rate_val in [(2, lr), (3, br), (4, fr), (5, mcdc)]:
                    if isinstance(rate_val, (int, float)):
                        fill = _coverage_fill(rate_val)
                        if fill:
                            ws.cell(row=row_idx, column=col_idx).fill = fill

        col_widths = [60, 14, 16, 18, 16]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"

    def _write_selftest_shall_coverage(self, wb: Workbook, review: dict):
        ws = wb.create_sheet("SHALL Coverage")
        ws.sheet_properties.tabColor = "8E44AD"

        headers = ["SHALL Statement", "Status", "Test Name", "Assertion Lines", "Auto Mapped"]
        ws.append(headers)
        _apply_header_style(ws, 1, len(headers))

        shall_statements = review.get("shall_statements", [])
        shall_to_tests_map = review.get("shall_auto_mapping", {})
        shall_assertion_map = review.get("shall_assertion_map", {})

        # If the review has shall_statements from spec, prefer those;
        # otherwise fall back to the auto-mapping keys
        if shall_statements:
            for shall in shall_statements:
                stmt = shall.get("statement", "")
                section = shall.get("section", "")
                line_num = shall.get("line", 0)
                display_stmt = f"[{section}] {stmt}" if section else stmt

                matched_tests = shall_to_tests_map.get(stmt, [])
                is_covered = len(matched_tests) > 0
                status = "✅" if is_covered else "❌"

                test_str = ", ".join(matched_tests[:3])
                if len(matched_tests) > 3:
                    test_str += f" (+{len(matched_tests) - 3} more)"

                # Assertion lines
                assertion_refs = shall_assertion_map.get(stmt, {})
                all_assertion_lines: list[int] = []
                for test_lines in assertion_refs.values():
                    all_assertion_lines.extend(test_lines)
                assertion_str = ", ".join(f"L{n}" for n in sorted(set(all_assertion_lines))[:10])
                if len(set(all_assertion_lines)) > 10:
                    assertion_str += " (+more)"

                auto_mapped = "yes" if is_covered else "no"

                row_data = [display_stmt, status, test_str, assertion_str, auto_mapped]
                ws.append(row_data)
                row_idx = ws.max_row
                _apply_body_style(ws, row_idx, len(headers))

                # Colour uncovered SHALLs
                if not is_covered:
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=col).fill = _FILL_RED_LIGHT

                # Hyperlink Test Name column (col C=3)
                if matched_tests:
                    cell_c = ws.cell(row=row_idx, column=3)
                    cell_c.font = _FONT_LINK
        else:
            # Fallback: use shall_auto_mapping keys directly
            if shall_to_tests_map:
                for stmt, test_names in sorted(shall_to_tests_map.items()):
                    is_covered = bool(test_names)
                    status = "✅" if is_covered else "❌"
                    test_str = ", ".join(test_names[:3])
                    if len(test_names) > 3:
                        test_str += f" (+{len(test_names) - 3} more)"

                    assertion_lines = shall_assertion_map.get(stmt, {})
                    all_lines = set()
                    for tl in assertion_lines.values():
                        all_lines.update(tl)
                    assertion_str = ", ".join(f"L{n}" for n in sorted(all_lines)[:10])
                    if len(all_lines) > 10:
                        assertion_str += " (+more)"

                    auto_mapped = "yes" if is_covered else "no"
                    row_data = [stmt[:120], status, test_str, assertion_str, auto_mapped]
                    ws.append(row_data)
                    row_idx = ws.max_row
                    _apply_body_style(ws, row_idx, len(headers))

                    if not is_covered:
                        for col in range(1, len(headers) + 1):
                            ws.cell(row=row_idx, column=col).fill = _FILL_RED_LIGHT

                    if test_names:
                        cell_c = ws.cell(row=row_idx, column=3)
                        cell_c.font = _FONT_LINK
            else:
                ws.append(["No SHALL coverage data available", "", "", "", ""])
                _apply_body_style(ws, 2, len(headers))

        col_widths = [60, 10, 40, 30, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"

    # ──────────────────────────────────────────────────────────────────────
    # Helpers
    # ──────────────────────────────────────────────────────────────────────

    @staticmethod
    def _get_tool_version() -> str:
        """Get cppcheck version string."""
        import subprocess
        try:
            result = subprocess.run(
                ["cppcheck", "--version"],
                capture_output=True, text=True, timeout=5,
            )
            return result.stdout.strip() or result.stderr.strip() or "unknown"
        except (FileNotFoundError, subprocess.TimeoutExpired, OSError):
            return "unknown"

    @staticmethod
    def _deviation_to_dict(dev) -> dict:
        """Normalize a deviation entry to a dict."""
        if isinstance(dev, dict):
            return {
                "deviation_rule": dev.get("rule_id", ""),
                "file_pattern": dev.get("file_pattern", ""),
                "reason": dev.get("reason", ""),
                "approved_by": dev.get("approved_by", ""),
                "risk_level": dev.get("risk_level", "mid"),
                "expires": dev.get("expires", ""),
                "status": dev.get("status", "pending"),
                "alm_ticket": dev.get("alm_ticket", ""),
            }
        if isinstance(dev, tuple):
            fields = tuple(dev)
            return {
                "deviation_rule": fields[0] if len(fields) > 0 else "",
                "file_pattern": fields[1] if len(fields) > 1 else "",
                "reason": fields[2] if len(fields) > 2 else "",
                "approved_by": fields[3] if len(fields) > 3 else "",
                "risk_level": fields[4] if len(fields) > 4 else "mid",
                "expires": fields[5] if len(fields) > 5 else "",
                "status": fields[6] if len(fields) > 6 else "pending",
                "alm_ticket": fields[7] if len(fields) > 7 else "",
            }
        return {
            "deviation_rule": getattr(dev, "rule_id", ""),
            "file_pattern": getattr(dev, "file_pattern", ""),
            "reason": getattr(dev, "reason", ""),
            "approved_by": getattr(dev, "approved_by", ""),
            "risk_level": getattr(dev, "risk_level", "mid"),
            "expires": getattr(dev, "expires", ""),
            "status": getattr(dev, "status", "pending"),
            "alm_ticket": getattr(dev, "alm_ticket", ""),
        }
