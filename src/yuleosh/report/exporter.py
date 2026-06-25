#!/usr/bin/env python3
# Copyright (c) 2025 frisky1985
# SPDX-License-Identifier: MIT

"""
CI Report Exporter — generates JSON, Markdown, and Excel reports from CI layer results.

Called automatically after each CI layer run to produce structured reports
under ``.yuleosh/reports/``.

Output files per layer:
  - layer{N}-report.json    — structured JSON
  - layer{N}-report.md      — formatted Markdown
  - layer{N}-report.xlsx    — Excel workbook (if openpyxl available)

Final report (after run_all):
  - ci-final-report.json
  - ci-final-report.md
  - ci-final-report.xlsx
"""

import json
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

log = logging.getLogger("report.exporter")

# ------------------------------------------------------------------
# Constants — avoid hardcoded path strings
# ------------------------------------------------------------------

_CI_RESULTS_DIR = ".osh/ci"
_REPORTS_DIR = ".yuleosh/reports"


def _load_ci_results(project_dir: str, layer: int) -> Optional[dict]:
    """Load the most recent CI result JSON for a given layer."""
    ci_dir = Path(project_dir) / _CI_RESULTS_DIR
    if not ci_dir.exists():
        return None
    prefix = f"layer{layer}-"
    candidates = sorted(
        [f for f in ci_dir.iterdir() if f.name.startswith(prefix) and f.suffix == ".json"],
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        return None
    try:
        return json.loads(candidates[0].read_text())
    except (json.JSONDecodeError, OSError) as e:
        log.warning("Cannot load CI result for layer %d: %s", layer, e)
        return None


def _load_misra_report(project_dir: str) -> Optional[dict]:
    """Load the latest MISRA report JSON if it exists."""
    path = Path(project_dir) / _REPORTS_DIR / "misra-report.json"
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as e:
        log.warning("Cannot load MISRA report: %s", e)
        return None


def _load_coverage_report(project_dir: str) -> Optional[dict]:
    """Load the latest C coverage JSON if it exists."""
    path = Path(project_dir) / _REPORTS_DIR / "c-coverage.json"
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text())
    except (json.JSONDecodeError, OSError):
        return None


def _serialize_layer_to_summary(results: dict) -> dict:
    """Extract a short summary from a layer result dict."""
    stages = results.get("stages", [])
    passed = sum(1 for s in stages if s.get("status") == "passed")
    failed = sum(1 for s in stages if s.get("status") == "failed")
    skipped = sum(1 for s in stages if s.get("status") == "skipped")
    coverage = results.get("coverage", {})
    return {
        "layer": results.get("layer"),
        "status": results.get("status", "unknown"),
        "commit": results.get("commit", ""),
        "started_at": results.get("started_at", ""),
        "completed_at": results.get("completed_at", ""),
        "stages_summary": {
            "total": len(stages),
            "passed": passed,
            "failed": failed,
            "skipped": skipped,
        },
        "stages": [
            {
                "name": s.get("name"),
                "status": s.get("status"),
                "detail": s.get("detail", ""),
            }
            for s in stages
        ],
        "coverage": {
            "line": coverage.get("line_coverage"),
            "condition": coverage.get("condition_coverage"),
            "line_pass": coverage.get("line_pass"),
            "condition_pass": coverage.get("condition_pass"),
        } if coverage else None,
        "errors": results.get("errors", []),
    }


def _collect_all_layers(project_dir: str) -> list[dict]:
    """Collect summaries from layers 1, 2, 2.5, 3."""
    layers = []
    for layer in [1, 2, 25, 3]:
        result = _load_ci_results(project_dir, layer)
        if result:
            layers.append(_serialize_layer_to_summary(result))
    return layers


# ------------------------------------------------------------------
# JSON Report
# ------------------------------------------------------------------


def generate_json_report(
    project_dir: str,
    layers_data: list[dict],
    misra: Optional[dict] = None,
    coverage: Optional[dict] = None,
    is_final: bool = False,
) -> dict:
    """Build the full JSON report dict."""
    report: dict = {
        "report_type": "final" if is_final else "incremental",
        "generated_at": datetime.now().isoformat(),
        "project_dir": project_dir,
        "layers": layers_data,
    }
    if misra:
        misra_summary = misra.get("summary", {})
        report["misra"] = {
            "total_violations": misra_summary.get("total_violations", 0),
            "required": misra_summary.get("misra_classification", {}).get("required", 0),
            "advisory": misra_summary.get("misra_classification", {}).get("advisory", 0),
            "violations_per_kloc": misra_summary.get("violations_per_kloc", 0),
            "rules_violated": misra_summary.get("total_rules_violated", 0),
            "files_affected": len(misra_summary.get("unique_files", [])),
        }
    if coverage:
        report["c_coverage"] = {
            "line_rate": coverage.get("line_rate", 0),
            "branch_rate": coverage.get("branch_rate", 0),
        }
    return report


# ------------------------------------------------------------------
# Markdown Report
# ------------------------------------------------------------------


def _status_emoji(status: str) -> str:
    _m = {"passed": "✅", "failed": "❌", "skipped": "⏭️", "warning": "⚠️", "error": "❌", "running": "⏳"}
    return _m.get(status, "❓")


def generate_markdown_report(
    project_dir: str,
    layers_data: list[dict],
    misra: Optional[dict] = None,
    coverage: Optional[dict] = None,
    is_final: bool = False,
) -> str:
    """Generate a Markdown CI report."""
    lines = [
        "# CI Report Summary",
        "",
        f"> Generated: {datetime.now().isoformat()}",
        f"> Project: {project_dir}",
        f"> Type: {'Final' if is_final else 'Incremental'}",
        "",
    ]

    # Overall status
    overall = "✅ ALL PASSED" if all(
        ld.get("status") == "passed" for ld in layers_data
    ) else "❌ FAILED"
    lines.append(f"## Overall: {overall}")
    lines.append("")

    # Per-layer
    lines.append("## Layer Summary")
    lines.append("")
    lines.append("| Layer | Status | Passed | Failed | Skipped | Errors |")
    lines.append("|:------|:-------|------:|------:|--------:|:-------|")
    for ld in layers_data:
        l = ld["layer"]
        emoji = _status_emoji(ld.get("status", ""))
        ss = ld.get("stages_summary", {})
        err_count = len(ld.get("errors", []))
        err_str = "⚠️" if err_count > 0 else "—"
        lines.append(
            f"| L{l} | {emoji} {ld.get('status', '?')} "
            f"| {ss.get('passed', 0)} | {ss.get('failed', 0)} "
            f"| {ss.get('skipped', 0)} | {err_str} |"
        )
    lines.append("")

    # Stage details
    for ld in layers_data:
        l = ld["layer"]
        stages = ld.get("stages", [])
        if not stages:
            continue
        lines.append(f"### Layer L{l} — Stage Details")
        lines.append("")
        lines.append("| Stage | Status | Detail |")
        lines.append("|:------|:-------|:-------|")
        for s in stages:
            lines.append(f"| {s.get('name', '?')} | {_status_emoji(s.get('status', ''))} | {s.get('detail', '')} |")
        lines.append("")

    # Coverage
    for ld in layers_data:
        cov = ld.get("coverage")
        if cov and cov.get("line") is not None:
            l = ld["layer"]
            line_emoji = "✅" if cov.get("line_pass") else "❌"
            cond_emoji = "✅" if cov.get("condition_pass") else "❌"
            lines.append(f"### Layer L{l} — Code Coverage")
            lines.append("")
            lines.append(f"| Metric | Value | Threshold | Status |")
            lines.append(f"|:-------|------:|---------:|:-------|")
            lines.append(f"| Line Coverage | {cov.get('line', 0):.1f}% | 85% | {line_emoji} |")
            lines.append(f"| Condition Coverage | {cov.get('condition', 0):.1f}% | 80% | {cond_emoji} |")
            lines.append("")

    # MISRA
    if misra:
        ms = misra.get("summary", {})
        totals = ms.get("misra_classification", {})
        lines.append("## MISRA C:2023")
        lines.append("")
        lines.append("| Metric | Value |")
        lines.append("|:-------|------:|")
        lines.append(f"| Total Violations | {ms.get('total_violations', 0)} |")
        lines.append(f"| Required | {totals.get('required', 0)} |")
        lines.append(f"| Advisory | {totals.get('advisory', 0)} |")
        lines.append(f"| Violations / KLOC | {ms.get('violations_per_kloc', 0)} |")
        lines.append(f"| Rules Violated | {ms.get('total_rules_violated', 0)} |")
        lines.append(f"| Files Affected | {len(ms.get('unique_files', []))} |")
        lines.append("")

    # C coverage
    if coverage:
        lines.append("## C/C++ Coverage")
        lines.append("")
        lines.append("| Metric | Value |")
        lines.append("|:-------|------:|")
        lines.append(f"| Line Rate | {coverage.get('line_rate', 0):.1f}% |")
        lines.append(f"| Branch Rate | {coverage.get('branch_rate', 0):.1f}% |")
        lines.append("")

    lines.append("---")
    lines.append("*Generated by yuleOSH CI Report Exporter*")
    return "\n".join(lines)


# ------------------------------------------------------------------
# Excel Report
# ------------------------------------------------------------------


def _write_excel_report(layers_data: list[dict], output_path: Path) -> None:
    """Write a formatted Excel report if openpyxl is available."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        log.warning("openpyxl not installed — skipping Excel report")
        return

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    normal_font = Font(size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    def _style_header(ws, cols: int):
        for c in range(1, cols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    headers = ["Layer", "Status", "Passed", "Failed", "Skipped", "Errors"]
    for c, h in enumerate(headers, 1):
        ws_summary.cell(row=1, column=c, value=h)
    _style_header(ws_summary, len(headers))
    for r, ld in enumerate(layers_data, 2):
        ss = ld.get("stages_summary", {})
        ws_summary.cell(row=r, column=1, value=f"L{ld['layer']}").font = normal_font
        ws_summary.cell(row=r, column=2, value=ld.get("status", "")).font = normal_font
        ws_summary.cell(row=r, column=3, value=ss.get("passed", 0)).font = normal_font
        ws_summary.cell(row=r, column=4, value=ss.get("failed", 0)).font = normal_font
        ws_summary.cell(row=r, column=5, value=ss.get("skipped", 0)).font = normal_font
        ws_summary.cell(row=r, column=6, value=", ".join(ld.get("errors", []))).font = normal_font
        for c in range(1, len(headers) + 1):
            ws_summary.cell(row=r, column=c).border = thin_border
    ws_summary.column_dimensions["A"].width = 10
    ws_summary.column_dimensions["B"].width = 12
    for col in "CDEF":
        ws_summary.column_dimensions[col].width = 10

    # Sheet 2: Stages
    ws_stages = wb.create_sheet("Stages")
    stage_headers = ["Layer", "Stage", "Status", "Detail"]
    for c, h in enumerate(stage_headers, 1):
        ws_stages.cell(row=1, column=c, value=h)
    _style_header(ws_stages, len(stage_headers))
    row = 2
    for ld in layers_data:
        for s in ld.get("stages", []):
            ws_stages.cell(row=row, column=1, value=f"L{ld['layer']}").font = normal_font
            ws_stages.cell(row=row, column=2, value=s.get("name", "")).font = normal_font
            ws_stages.cell(row=row, column=3, value=s.get("status", "")).font = normal_font
            ws_stages.cell(row=row, column=4, value=s.get("detail", "")).font = normal_font
            for c in range(1, len(stage_headers) + 1):
                ws_stages.cell(row=row, column=c).border = thin_border
                ws_stages.cell(row=row, column=c).alignment = wrap
            row += 1
    ws_stages.column_dimensions["A"].width = 8
    ws_stages.column_dimensions["B"].width = 22
    ws_stages.column_dimensions["C"].width = 12
    ws_stages.column_dimensions["D"].width = 50

    wb.save(str(output_path))
    log.info("Excel report saved: %s", output_path)


# ------------------------------------------------------------------
# Public API
# ------------------------------------------------------------------


def generate_layer_report(project_dir: str, layer: int) -> Optional[Path]:
    """Generate report files for a single CI layer run.

    Reads the layer result from .osh/ci/ and writes to .yuleosh/reports/.

    Returns the report output directory path, or None on failure.
    """
    result = _load_ci_results(project_dir, layer)
    if not result:
        log.warning("No CI result found for layer %d in %s", layer, project_dir)
        return None

    report_dir = Path(project_dir) / _REPORTS_DIR
    report_dir.mkdir(parents=True, exist_ok=True)

    summary = _serialize_layer_to_summary(result)
    misra = _load_misra_report(project_dir)
    coverage = _load_coverage_report(project_dir)

    layer_label = str(layer).replace(".", "_")
    prefix = f"layer{layer_label}"

    # JSON
    json_data = generate_json_report(project_dir, [summary], misra, coverage, is_final=False)
    json_path = report_dir / f"{prefix}-report.json"
    json_path.write_text(json.dumps(json_data, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    log.info("Layer %d JSON report: %s", layer, json_path)

    # Markdown
    md_content = generate_markdown_report(project_dir, [summary], misra, coverage, is_final=False)
    md_path = report_dir / f"{prefix}-report.md"
    md_path.write_text(md_content, encoding="utf-8")
    log.info("Layer %d MD report: %s", layer, md_path)

    # Excel
    excel_path = report_dir / f"{prefix}-report.xlsx"
    _write_excel_report([summary], excel_path)

    return report_dir


def generate_final_report(project_dir: str) -> Optional[Path]:
    """Generate the final comprehensive CI report after run_all.

    Collects all layer results and writes to .yuleosh/reports/ci-final-report.*.

    Returns the report output directory path, or None on failure.
    """
    report_dir = Path(project_dir) / _REPORTS_DIR
    report_dir.mkdir(parents=True, exist_ok=True)

    layers_data = _collect_all_layers(project_dir)
    if not layers_data:
        log.warning("No CI layer results found in %s", project_dir)
        return None

    misra = _load_misra_report(project_dir)
    coverage = _load_coverage_report(project_dir)

    # JSON
    json_data = generate_json_report(project_dir, layers_data, misra, coverage, is_final=True)
    json_path = report_dir / "ci-final-report.json"
    json_path.write_text(json.dumps(json_data, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    log.info("Final JSON report: %s", json_path)

    # Markdown
    md_content = generate_markdown_report(project_dir, layers_data, misra, coverage, is_final=True)
    md_path = report_dir / "ci-final-report.md"
    md_path.write_text(md_content, encoding="utf-8")
    log.info("Final MD report: %s", md_path)

    # Excel
    excel_path = report_dir / "ci-final-report.xlsx"
    _write_excel_report(layers_data, excel_path)

    # Also generate per-layer reports for completeness
    for ld in layers_data:
        l = ld["layer"]
        generate_layer_report(project_dir, l)

    # 飞书 Webhook 自动推送（如果设置了 FEISHU_WEBHOOK_URL）
    _auto_feishu_notify(project_dir)

    return report_dir


# ------------------------------------------------------------------
# Feishu Webhook Auto-notify
# ------------------------------------------------------------------


def _auto_feishu_notify(project_dir: str) -> None:
    """如果环境变量 FEISHU_WEBHOOK_URL 已设置，自动推送质量卡片到飞书。"""
    feishu_url = os.environ.get("FEISHU_WEBHOOK_URL", "").strip()
    if not feishu_url:
        return

    try:
        from yuleosh.report.feishu_notifier import post_quality_card_to_feishu
        success = post_quality_card_to_feishu(
            webhook_url=feishu_url,
            project_dir=project_dir,
        )
        if success:
            log.info("Feishu quality card auto-pushed successfully")
        else:
            log.warning("Feishu quality card auto-push failed")
    except ImportError:
        log.debug("feishu_notifier not available — skipping auto-push")
    except Exception as e:
        log.warning("Feishu auto-push error: %s", e)
